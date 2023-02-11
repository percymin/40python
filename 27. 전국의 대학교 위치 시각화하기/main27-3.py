# 라이브러리를 불러온다
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

#판다스를 이용하여 엑셀파일을 불러온다
filePath = r'27. 전국의 대학교 위치 시각화하기\고등교육기관 하반기 주소록(2022).xlsx'
df_from_excel = pd.read_excel(filePath,engine='openpyxl')
# 5개줄부터 columns로 지정한다
df_from_excel.columns = df_from_excel.loc[4].tolist()
#0~5번째 줄은 삭제 불필요한 정보
df_from_excel = df_from_excel.drop(index=list(range(0,5)))

#주소를 좌표로 변환하는 api를 사용하는 함수를 만들었다
url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'   #도로명주소
road_type2 = 'PARCEL' #지번주소
address = '&address='
keys = '&key='
primary_key = '693598C2-0E00-3498-B514-EAD0B01C5CD5'

def request_geo(road):
    #위에서 정의한 정보들 집어넣기
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        # status ok일때 x y에 정보 넣기
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x,y
    else:
        # status ok아닐때는 xy둘다 0으로 리턴한다
        x = 0
        y = 0
        return x,y

#주소를 좌표로 바꿔 학교주소좌표.xlsx 엑셀파일을 생성후 저장
try:
    wb = load_workbook(r"27. 전국의 대학교 위치 시각화하기\학교주소좌표.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

#정보 입력
university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()


for num,value in enumerate(address_list):
	#주소에서 괋를 삭제한다
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)
    #api를 활용하여 주소를 좌표로 변환한다
    x,y = request_geo(addr)
    #학교명 순서대로 엑셀에 저장
    sheet.append([university_list[num],addr,x,y])


wb.save(r"27. 전국의 대학교 위치 시각화하기\학교주소좌표.xlsx")