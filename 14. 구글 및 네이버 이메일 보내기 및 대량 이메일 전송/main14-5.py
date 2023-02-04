from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
#메일을 이용하기위한 라이브러리를 이용한다.

load_wb = load_workbook(r"14. 구글 및 네이버 이메일 보내기 및 대량 이메일 전송\이메일주소.xlsx", data_only=True)
load_ws =load_wb.active
#파일 14->이메일주소.xlsx를 읽어온다.

for i in range(1,load_ws.max_row + 1): #읽어온 이메일주소.xlsx 파일에서 반복문을 돌린다.
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:",recv_email_value)
    try:
        send_email = "sammymin@naver.com"
        send_pwd = "awds4818~"

        recv_email = recv_email_value # 위에 for반복문으로 돌리는 이메일 주소를 받는 사람 이메일에 저장한다.

        smtp_name = "smtp.naver.com" 
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] ="엑셀에서 메일주소를 읽어 자동으로 보내는 메일입니다."
        msg['From'] = send_email 
        msg['To'] = recv_email 

        text =  """
                메일내용 입니다.
                감사합니다.
                """

        msg.attach( MIMEText(text) ) 

        s=smtplib.SMTP( smtp_name , smtp_port )
        s.starttls()
        s.login( send_email , send_pwd )
        s.sendmail( send_email, recv_email, msg.as_string() )
        s.quit()
		#나머지는 메일 보내는 구성요소, 메일 보내기이다
    except:
        print("에러:",recv_email_value)
	# 이메일 주소에 에러가 있을시 except로 에러라고 표시한다.